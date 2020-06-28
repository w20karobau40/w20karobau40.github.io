import json

import Levenshtein
from loguru import logger
from openpyxl import load_workbook


@logger.catch
def main():
    HEADER_ROW = 3
    MAX_ROW = 37
    COLS = 90
    with open('question_and_answers.json') as input_file:
        question_and_answers = json.load(input_file)
    questions = []
    for qa in question_and_answers:
        questions.extend(qa['questions'])
    wb = load_workbook('../karobau40/200206_Ergebnisse_IWU_Karosseriebau.xlsx')
    ws = wb.active
    users = {row: {'id': str(row)} for row in range(HEADER_ROW + 1, MAX_ROW + 1)}
    for col in range(1, COLS + 1):
        # first get correct question / subquestion
        optimum = 10 ** 6
        question_index, subquestion_index = -1, -1
        header = ws.cell(HEADER_ROW, col).value
        if col == 47:
            wait = 1
        for q_i, q in enumerate(questions):
            if q_i == 13:
                wait = 2
            answeroptions = q.get('answeroptions', [{'text': 'not quoted'}, {'text': 'quoted'}])
            answer_distance = sum(min(Levenshtein.distance(str(ws.cell(row, col).value), ao['text']) for ao in answeroptions) for row in range(HEADER_ROW + 1, MAX_ROW + 1)) / (MAX_ROW - HEADER_ROW)
            if 'subquestions' in q.keys():
                for sq_i, sq in enumerate(q['subquestions']):
                    text_distance = Levenshtein.distance(header, sq['text'])
                    if answer_distance ** 2 + text_distance ** 2 < optimum:
                        optimum = answer_distance ** 2 + text_distance ** 2
                        question_index, subquestion_index = q_i, sq_i
            else:
                text_distance = Levenshtein.distance(header, q['text'])
                if answer_distance ** 2 + text_distance ** 2 < optimum:
                    optimum = answer_distance ** 2 + text_distance ** 2
                    question_index, subquestion_index = q_i, -1
        if optimum > 10:
            # ignore text questions
            continue
        question = questions[question_index]
        if subquestion_index > -1:
            code = f"{question['code']}[{question['subquestions'][subquestion_index]['code']}]"
        else:
            code = question['code']
        answeroptions = question.get('answeroptions', [{'text': 'not quoted', 'code': ''}, {'text': 'quoted', 'code': 'Y'}])
        for row in range(HEADER_ROW + 1, MAX_ROW + 1):
            if ws.cell(row, col).value == 0:
                users[row][code] = ""
            else:
                option = min(answeroptions, key=lambda ao: Levenshtein.distance(ao['text'], str(ws.cell(row, col).value)))
                users[row][code] = option['code']

    responses = [{str(key): value} for key, value in users.items()]
    with open('old_responses.json', 'w') as output:
        json.dump({"responses": responses}, output)


if __name__ == '__main__':
    main()
